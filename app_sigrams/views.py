from datetime import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from app_registros.models import UserProfile, Productor, MarcaSenal, Solicitud, Campo, TipoSenal, ImagenMarcaPredefinida
from app_registros.forms import ProductorForm, MarcaSenalForm, SolicitudForm
from django.http import JsonResponse
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy, reverse
from django.conf import settings

from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView, TemplateView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from app_registros.models import Productor, Campo, MarcaSenal


from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.http import HttpResponse


import datetime
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth

from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create your views here.

from django.http import JsonResponse
import json

def get_campos_por_productor(request, productor_id):
    """Obtener campos de un productor específico (para AJAX)"""
    try:
        # Verificar si el productor existe
        from app_registros.models import Productor, Campo
        productor = Productor.objects.get(id=productor_id)
        
        # Obtener campos del productor
        campos = Campo.objects.filter(productor_id=productor_id)
        
        if not campos.exists():
            # Si no hay campos, crear uno automáticamente basado en los datos del productor
            campo = Campo.objects.create(
                nombre=productor.campo or f"Campo de {productor.nombre} {productor.apellido}",
                productor=productor,
                distrito=productor.localidad or "Sin especificar",
                departamento=productor.departamento or "Sin especificar",
                area_hectareas=productor.area_hectareas or 0,
                latitud=productor.latitud,
                longitud=productor.longitud
            )
            campos = Campo.objects.filter(productor_id=productor_id)
        
        # Preparar datos para JSON
        data = []
        for campo in campos:
            data.append({
                'id': campo.id,
                'nombre': campo.nombre,
                'distrito': campo.distrito or '',
                'departamento': campo.departamento or ''
            })
        
        return JsonResponse(data, safe=False)
        
    except Productor.DoesNotExist:
        return JsonResponse([], safe=False)
    except Exception as e:
        print(f"Error en get_campos_por_productor: {str(e)}")
        return JsonResponse([{'error': str(e)}], safe=False, status=500)
    
def get_imagenes_marcas(request):
    """Obtener imágenes predefinidas de marcas (para AJAX)"""
    try:
        imagenes = ImagenMarcaPredefinida.objects.filter(activa=True)
        data = []
        for imagen in imagenes:
            data.append({
                'id': imagen.id,
                'nombre': imagen.nombre,
                'tipo_marca': imagen.get_tipo_marca_display(),
                'imagen': imagen.imagen.url if imagen.imagen else '',
                'imagen_url': request.build_absolute_uri(imagen.imagen.url) if imagen.imagen else ''
            })
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


from django.http import JsonResponse
from app_registros.models import MarcaSenal, Productor
import json

def get_marcas_por_productor(request, productor_id):
    """Obtener marcas de un productor específico (para AJAX)"""
    try:
        print(f"DEBUG SERVIDOR: Solicitando marcas para productor_id={productor_id}")
        
        # Verificar que el productor existe
        if not Productor.objects.filter(id=productor_id).exists():
            return JsonResponse([], safe=False)
        
        # Obtener marcas del productor
        marcas = MarcaSenal.objects.filter(productor_id=productor_id).order_by('-fecha_inscripcion')
        
        # Preparar datos para JSON
        data = []
        for marca in marcas:
            data.append({
                'id': marca.id,
                'numero_orden': marca.numero_orden,
                'descripcion_marca': marca.descripcion_marca,
                'tipo_tramite': marca.get_tipo_tramite_display(),
                'estado': marca.get_estado_display(),
                'fecha_inscripcion': marca.fecha_inscripcion.strftime('%d/%m/%Y') if marca.fecha_inscripcion else 'No especificada',
            })
        
        print(f"DEBUG SERVIDOR: Enviando {len(data)} marcas")
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        print(f"ERROR SERVIDOR en get_marcas_por_productor: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse([{'error': str(e)}], safe=False, status=500)


# app_sigrams/views.py - Actualizar función home
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Count, Sum
import datetime
import json


@login_required
def home(request):
    """Vista principal del dashboard"""
    from app_registros.models import Productor, MarcaSenal, Solicitud, Campo

    hoy = timezone.now()

    # ========== ESTADÍSTICAS BÁSICAS ==========
    total_productores = Productor.objects.count()
    total_marcas = MarcaSenal.objects.filter(estado='VIGENTE').count()
    solicitudes_pendientes = Solicitud.objects.filter(estado='PENDIENTE').count()
    total_campos = Campo.objects.count()

    # ========== INGRESOS ==========
    # Primer día del mes actual
    primer_dia_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Ingresos mes actual
    ingresos_mes = (
        MarcaSenal.objects.filter(
            fecha_creacion__gte=primer_dia_mes,
            valor_sellado__isnull=False
        ).aggregate(total=Sum('valor_sellado'))['total']
        or 0
    )

    # Ingresos mes anterior
    primer_dia_mes_anterior = (primer_dia_mes - datetime.timedelta(days=1)).replace(day=1)
    ultimo_dia_mes_anterior = primer_dia_mes - datetime.timedelta(days=1)

    ingresos_mes_anterior = (
        MarcaSenal.objects.filter(
            fecha_creacion__gte=primer_dia_mes_anterior,
            fecha_creacion__lte=ultimo_dia_mes_anterior,
            valor_sellado__isnull=False
        ).aggregate(total=Sum('valor_sellado'))['total']
        or 0
    )

    # Variación porcentual
    if ingresos_mes_anterior > 0:
        variacion_ingresos = ((ingresos_mes - ingresos_mes_anterior) / ingresos_mes_anterior) * 100
    else:
        variacion_ingresos = 100 if ingresos_mes > 0 else 0

    # ========== TRÁMITES ==========
    tramites_mes = Solicitud.objects.filter(
        fecha_solicitud__gte=primer_dia_mes
    ).count()

    # ========== DISTRIBUCIONES ==========
    productores_por_estado = list(
        Productor.objects.values('estado')
        .annotate(total=Count('id'))
        .order_by('estado')
    )

    marcas_por_tipo = list(
        MarcaSenal.objects.values('tipo_tramite')
        .annotate(total=Count('id'))
        .order_by('tipo_tramite')
    )

    solicitudes_por_estado = list(
        Solicitud.objects.values('estado')
        .annotate(total=Count('id'))
        .order_by('estado')
    )

    # ========== DATOS RECIENTES ==========
    ultimas_solicitudes = (
        Solicitud.objects.select_related('productor')
        .order_by('-fecha_solicitud')[:10]
    )

    fecha_limite = hoy - datetime.timedelta(days=7)
    productores_recientes = (
        Productor.objects.filter(fecha_registro__gte=fecha_limite)
        .order_by('-fecha_registro')[:5]
    )

    fecha_vencimiento = hoy + datetime.timedelta(days=30)
    marcas_por_vencer = (
        MarcaSenal.objects.filter(
            fecha_vencimiento__gte=hoy,
            fecha_vencimiento__lte=fecha_vencimiento,
            estado='VIGENTE'
        )
        .select_related('productor')
        .order_by('fecha_vencimiento')[:5]
    )

    marcas_recientes = (
        MarcaSenal.objects.select_related('productor')
        .order_by('-fecha_inscripcion')[:5]
    )

    # ========== ACTIVIDAD MENSUAL (ÚLTIMOS 6 MESES) ==========
    meses_actividad = []

    for i in range(6):
        fecha = hoy - datetime.timedelta(days=30 * i)
        mes_str = fecha.strftime('%Y-%m')
        nombre_mes = fecha.strftime('%b')

        solicitudes_mes = Solicitud.objects.filter(
            fecha_solicitud__year=fecha.year,
            fecha_solicitud__month=fecha.month
        ).count()

        marcas_mes = MarcaSenal.objects.filter(
            fecha_inscripcion__year=fecha.year,
            fecha_inscripcion__month=fecha.month
        ).count()

        ingresos_mes_valor = (
            MarcaSenal.objects.filter(
                fecha_creacion__year=fecha.year,
                fecha_creacion__month=fecha.month,
                valor_sellado__isnull=False
            ).aggregate(total=Sum('valor_sellado'))['total']
            or 0
        )

        meses_actividad.append({
            'mes': mes_str,
            'nombre': nombre_mes,
            'solicitudes': solicitudes_mes,
            'marcas': marcas_mes,
            'ingresos': float(ingresos_mes_valor)
        })

    meses_actividad.reverse()

    # ========== TOP PRODUCTORES ==========
    top_productores = (
        Productor.objects.annotate(total_marcas=Count('marcas_senales'))
        .order_by('-total_marcas')[:5]
    )

    # ========== CONTEXT ==========
    context = {
        # Básicos
        'total_productores': total_productores,
        'total_marcas': total_marcas,
        'solicitudes_pendientes': solicitudes_pendientes,
        'total_campos': total_campos,

        # Finanzas
        'ingresos_mes': ingresos_mes,
        'ingresos_mes_anterior': ingresos_mes_anterior,
        'variacion_ingresos': variacion_ingresos,
        'tramites_mes': tramites_mes,

        # Gráficos
        'productores_por_estado': json.dumps(productores_por_estado),
        'marcas_por_tipo': json.dumps(marcas_por_tipo),
        'solicitudes_por_estado': json.dumps(solicitudes_por_estado),
        'meses_actividad': json.dumps(meses_actividad),

        # Listados
        'ultimas_solicitudes': ultimas_solicitudes,
        'productores_recientes': productores_recientes,
        'marcas_por_vencer': marcas_por_vencer,
        'marcas_recientes': marcas_recientes,
        'top_productores': top_productores,

        # Badges / listas
        'productores_estados_list': productores_por_estado,
        'marcas_tipos_list': marcas_por_tipo,
    }

    return render(request, 'app_sigrams/index.html', context)


from django.db.models.functions import TruncMonth
from django.db.models import Count, Sum
from django.utils import timezone
from datetime import timedelta

@login_required
def dashboard_admin(request):
    user_profile = UserProfile.objects.get(user=request.user)
    if user_profile.rol != 'admin':
        messages.error(request, 'No tiene permisos para acceder.')
        return redirect('home')

    # Usuarios
    total_usuarios = User.objects.count()
    admins = UserProfile.objects.filter(rol='admin').count()
    empleados = UserProfile.objects.filter(rol='empleado').count()

    # Solicitudes
    total_solicitudes = Solicitud.objects.count()
    solicitudes_pendientes = Solicitud.objects.filter(estado='PENDIENTE').count()
    solicitudes_en_revision = Solicitud.objects.filter(estado='EN_REVISION').count()
    solicitudes_aprobadas = Solicitud.objects.filter(estado='APROBADO').count()
    solicitudes_rechazadas = Solicitud.objects.filter(estado='RECHAZADO').count()

    # Productores y Marcas
    total_productores = Productor.objects.count()
    total_marcas = MarcaSenal.objects.count()

    # Ingresos (últimos 12 meses)
    hoy = timezone.now()
    doce_meses = hoy - timedelta(days=365)
    ingresos_por_mes = (
        MarcaSenal.objects
        .filter(fecha_creacion__gte=doce_meses, valor_sellado__isnull=False)
        .annotate(mes=TruncMonth('fecha_creacion'))
        .values('mes')
        .annotate(total=Sum('valor_sellado'))
        .order_by('mes')
    )

    # Logs recientes (últimos 20)
    logs_recientes = ChangeLog.objects.select_related('user').order_by('-timestamp')[:20]

    # Últimas solicitudes
    ultimas_solicitudes = Solicitud.objects.select_related('productor', 'solicitante').order_by('-fecha_solicitud')[:10]

    context = {
        'total_usuarios': total_usuarios,
        'admins': admins,
        'empleados': empleados,
        'total_solicitudes': total_solicitudes,
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_en_revision': solicitudes_en_revision,
        'solicitudes_aprobadas': solicitudes_aprobadas,
        'solicitudes_rechazadas': solicitudes_rechazadas,
        'total_productores': total_productores,
        'total_marcas': total_marcas,
        'ingresos_por_mes': list(ingresos_por_mes),
        'logs_recientes': logs_recientes,
        'ultimas_solicitudes': ultimas_solicitudes,
    }
    return render(request, 'app_sigrams/admin/dashboard.html', context)



# LOGIN
def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
        
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'¡Bienvenido/a {user.username}!')
            
            # Redirigir según el rol del usuario
            try:
                profile = UserProfile.objects.get(user=user)
                if profile.rol == 'admin':
                    return redirect('home')
                elif profile.rol == 'inspector':
                    return redirect('home')
                else:
                    return redirect('home')
            except UserProfile.DoesNotExist:
                return redirect('home')
                
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    
    return render(request, 'app_sigrams/login.html')


# LOGOUT
@login_required
def logout_view(request):
    logout(request)
    messages.info(request, 'Has cerrado sesión correctamente.')
    return redirect('login')


# REGISTRO
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User

from app_sigrams.forms import CustomUserCreationForm

def register_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)

        if form.is_valid():
            user = form.save()

            # Crear o obtener perfil (evita duplicados)
            profile, created = UserProfile.objects.get_or_create(
                user=user,
                defaults={'rol': 'empleado'}
            )
            if not created:
                # Si ya existía (caso raro), aseguramos que tenga rol empleado
                profile.rol = 'empleado'
                profile.save()

            messages.success(
                request,
                'Cuenta creada correctamente. Ahora podés iniciar sesión.'
            )
            return redirect('login')

        else:
            # Mostrar errores campo por campo como mensajes
            for field, errors in form.errors.items():
                label = (
                    form.fields[field].label
                    if field in form.fields
                    else field
                )
                for error in errors:
                    messages.error(request, f"{label}: {error}")

    else:
        form = CustomUserCreationForm()

    return render(
        request,
        'app_sigrams/register.html',
        {'form': form}
    )



class EsEmpleadoOMas(UserPassesTestMixin):
    def test_func(self):
        user_profile = UserProfile.objects.get(user=self.request.user)
        return user_profile.rol in ['admin', 'empleado']



# ============================================================================
# VISTAS PARA PRODUCTORES
# ============================================================================

@login_required
def lista_productores(request):
    """Lista todos los productores con filtros"""
    productores = Productor.objects.all()
    
    # Filtros - ELIMINAMOS DISTRITO
    query = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    localidad = request.GET.get('localidad', '')
    departamento = request.GET.get('departamento', '')
    
    if query:
        productores = productores.filter(
            Q(nombre__icontains=query) | 
            Q(apellido__icontains=query) |
            Q(dni__icontains=query) |
            Q(campo__icontains=query)
        )
    
    if estado:
        productores = productores.filter(estado=estado)
    
    if localidad:
        productores = productores.filter(localidad__icontains=localidad)
    
    if departamento:
        productores = productores.filter(departamento__icontains=departamento)
    
    context = {
        'productores': productores,
        'query': query,
        'estado_filtro': estado,
        'localidad_filtro': localidad,
        'departamento_filtro': departamento,
    }
    return render(request, 'app_sigrams/productores/lista.html', context)

@login_required
def detalle_productor(request, pk):
    """Detalle de un productor específico"""
    productor = get_object_or_404(Productor, pk=pk)
    marcas = productor.marcas_senales.all()
    solicitudes = productor.solicitudes.all()
    campos = productor.campos.all()
    
    context = {
        'productor': productor,
        'marcas': marcas,
        'solicitudes': solicitudes,
        'campos': campos,
    }
    return render(request, 'app_sigrams/productores/detalle.html', context)

@login_required
def nuevo_productor(request):
    """Crear nuevo productor"""
    try:
        if request.method == 'POST':
            form = ProductorForm(request.POST)
            if form.is_valid():
                productor = form.save()
                messages.success(request, f'Productor {productor.nombre_completo} creado exitosamente.')
                return redirect('detalle_productor', pk=productor.pk)
            else:
                # Obtener todos los errores para mostrarlos
                error_list = []
                for field, errors in form.errors.items():
                    for error in errors:
                        if field != '__all__':
                            field_label = form.fields[field].label or field
                            error_list.append(f"{field_label}: {error}")
                        else:
                            error_list.append(str(error))
                
                if error_list:
                    messages.error(request, "Errores en el formulario:")
                    for error in error_list:
                        messages.error(request, f"- {error}")
                else:
                    messages.error(request, 'Por favor, corrija los errores en el formulario.')
                
                # Mantener los datos ingresados
                context = {'form': form, 'titulo': 'Nuevo Productor'}
                return render(request, 'app_sigrams/productores/form.html', context)
        else:
            form = ProductorForm()
        
        context = {'form': form, 'titulo': 'Nuevo Productor'}
        return render(request, 'app_sigrams/productores/form.html', context)
    
    except Exception as e:
        messages.error(request, f'Error al crear el productor: {str(e)}')
        return redirect('lista_productores')



@login_required
def editar_productor(request, pk):
    """Editar productor existente"""
    try:
        productor = get_object_or_404(Productor, pk=pk)
        
        if request.method == 'POST':
            form = ProductorForm(request.POST, instance=productor)
            if form.is_valid():
                productor = form.save()
                messages.success(request, f'Productor {productor.nombre_completo} actualizado exitosamente.')
                return redirect('detalle_productor', pk=productor.pk)
            else:
                messages.error(request, 'Por favor, corrija los errores en el formulario.')
        else:
            form = ProductorForm(instance=productor)
        
        context = {'form': form, 'titulo': 'Editar Productor', 'productor': productor}
        return render(request, 'app_sigrams/productores/form.html', context)
    
    except Exception as e:
        messages.error(request, f'Error al editar el productor: {str(e)}')
        return redirect('lista_productores')



@login_required
def eliminar_productor(request, pk):
    """Eliminar productor"""
    productor = get_object_or_404(Productor, pk=pk)
    
    if request.method == 'POST':
        nombre = productor.nombre_completo
        productor.delete()
        messages.success(request, f'Productor {nombre} eliminado exitosamente.')
        return redirect('lista_productores')
    
    context = {'productor': productor}
    return render(request, 'app_sigrams/productores/confirmar_eliminar.html', context)

# Vistas para Productores
class ListaProductoresView(LoginRequiredMixin, ListView):
    model = Productor
    template_name = 'productores/lista.html'
    context_object_name = 'productores'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Filtros
        nombre = self.request.GET.get('nombre')
        dni = self.request.GET.get('dni')
        localidad = self.request.GET.get('localidad')
        estado = self.request.GET.get('estado')
        
        if nombre:
            queryset = queryset.filter(Q(nombre__icontains=nombre) | Q(apellido__icontains=nombre))
        if dni:
            queryset = queryset.filter(dni__icontains=dni)
        if localidad:
            queryset = queryset.filter(localidad__icontains=localidad)
        if estado:
            queryset = queryset.filter(estado=estado)
        
        return queryset.order_by('apellido', 'nombre')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_productores'] = Productor.objects.count()
        context['estados'] = Productor.ESTADO_CHOICES
        return context

class NuevoProductorView(LoginRequiredMixin, EsEmpleadoOMas, CreateView):
    model = Productor
    template_name = 'productores/form.html'
    fields = ['nombre', 'apellido', 'dni', 'cuit', 'domicilio', 'distrito', 
              'localidad', 'telefono', 'email', 'latitud', 'longitud', 
              'estado', 'observaciones']
    success_url = reverse_lazy('lista_productores')
    
    def form_valid(self, form):
        form.instance.usuario_registro = self.request.user
        return super().form_valid(form)

class EditarProductorView(LoginRequiredMixin, EsEmpleadoOMas, UpdateView):
    model = Productor
    template_name = 'productores/form.html'
    fields = ['nombre', 'apellido', 'dni', 'cuit', 'domicilio', 'distrito', 
              'localidad', 'telefono', 'email', 'latitud', 'longitud', 
              'estado', 'observaciones']
    success_url = reverse_lazy('lista_productores')

class DetalleProductorView(LoginRequiredMixin, DetailView):
    model = Productor
    template_name = 'productores/detalle.html'
    context_object_name = 'productor'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['campos'] = self.object.campos.all()
        context['marcas'] = self.object.marcas_senales.all()
        context['solicitudes'] = self.object.solicitudes.all()
        return context

class EliminarProductorView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Productor
    template_name = 'productores/confirmar_eliminar.html'
    success_url = reverse_lazy('lista_productores')
    
    def test_func(self):
        user_profile = UserProfile.objects.get(user=self.request.user)
        return user_profile.rol == 'admin'


class ReporteProductoresPDFView(LoginRequiredMixin, View):
    def get(self, request):
        # Crear respuesta HTTP con contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_productores.pdf"'
        
        # Crear documento PDF
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph("Reporte de Productores - SIGRAMS", styles['Title'])
        elements.append(title)
        
        # Fecha
        fecha = Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
        elements.append(fecha)
        
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
        # Datos
        productores = Productor.objects.all().order_by('apellido', 'nombre')
        
        # Tabla
        data = [['Nombre', 'DNI', 'Localidad', 'Estado', 'Fecha Registro']]
        
        for p in productores:
            data.append([
                f"{p.apellido}, {p.nombre}",
                p.dni,
                p.localidad,
                p.get_estado_display(),
                p.fecha_registro.strftime('%d/%m/%Y')
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0A2E5A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        
        # Generar PDF
        doc.build(elements)
        
        return response

# ============================================================================
# VISTAS PARA MARCAS Y SEÑALES
# ============================================================================

# Vistas para Marcas y Señales
@login_required
def lista_marcas(request):
    """Lista todas las marcas y señales"""
    marcas = MarcaSenal.objects.select_related('productor', 'campo').all()
    
    # Filtros
    query = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    tipo_tramite = request.GET.get('tipo_tramite', '')
    
    if query:
        marcas = marcas.filter(
            Q(productor__nombre__icontains=query) | 
            Q(productor__apellido__icontains=query) |
            Q(numero_orden__icontains=query)
        )
    
    if estado:
        marcas = marcas.filter(estado=estado)
    
    if tipo_tramite:
        marcas = marcas.filter(tipo_tramite=tipo_tramite)
    
    context = {
        'marcas': marcas,
        'query': query,
        'estado_filtro': estado,
        'tipo_tramite_filtro': tipo_tramite,
    }
    return render(request, 'app_sigrams/marcas/lista.html', context)



@login_required
def detalle_marca(request, pk):
    """Detalle de una marca específica"""
    marca = get_object_or_404(MarcaSenal.objects.select_related('productor', 'campo', 'tipo_senal'), pk=pk)
    
    context = {
        'marca': marca,
    }
    return render(request, 'app_sigrams/marcas/detalle.html', context)



@login_required
def nueva_marca(request):
    """Crear nueva marca y señal"""

    if request.method == 'POST':
        form = MarcaSenalForm(request.POST, request.FILES)

        if form.is_valid():
            marca = form.save(commit=False)

            imagenes_predefinidas = form.cleaned_data.get('imagenes_predefinidas')

            # Si no subieron imagen manual pero eligieron predefinidas
            if imagenes_predefinidas and not request.FILES.get('imagen_marca'):
                primera_imagen = imagenes_predefinidas.first()

                if primera_imagen and primera_imagen.imagen:
                    contenido = primera_imagen.imagen.read()
                    nombre_archivo = os.path.basename(primera_imagen.imagen.name)

                    marca.imagen_marca.save(
                        f"predef_{primera_imagen.id}_{nombre_archivo}",
                        ContentFile(contenido),
                        save=False
                    )

            # 🔥 Aquí se genera automáticamente numero_orden
            marca.save()

            form.save_m2m()

            messages.success(
                request,
                f'Marca #{marca.numero_orden} creada exitosamente.'
            )

            return redirect('detalle_marca', pk=marca.pk)

        messages.error(request, 'Por favor, corrija los errores en el formulario.')

    else:
        form = MarcaSenalForm()

    return render(request, 'app_sigrams/marcas/form.html', {
        'form': form,
        'titulo': 'Nueva Marca y Señal',
        'imagenes_predefinidas': ImagenMarcaPredefinida.objects.filter(activa=True)
    })




@login_required
def editar_marca(request, pk):
    """Editar marca existente"""

    marca = get_object_or_404(MarcaSenal, pk=pk)

    if request.method == 'POST':
        form = MarcaSenalForm(request.POST, request.FILES, instance=marca)

        if form.is_valid():
            marca_temp = form.save(commit=False)

            imagenes_predefinidas = form.cleaned_data.get('imagenes_predefinidas')

            if (
                imagenes_predefinidas
                and not request.FILES.get('imagen_marca')
                and not marca.imagen_marca
            ):
                primera_imagen = imagenes_predefinidas.first()

                if primera_imagen and primera_imagen.imagen:
                    contenido = primera_imagen.imagen.read()
                    nombre_archivo = os.path.basename(primera_imagen.imagen.name)

                    marca_temp.imagen_marca.save(
                        f"predef_{primera_imagen.id}_{nombre_archivo}",
                        ContentFile(contenido),
                        save=False
                    )

            # 🔥 No modifica numero_orden porque ya existe
            marca_temp.save()

            form.save_m2m()

            messages.success(
                request,
                f'Marca #{marca_temp.numero_orden} actualizada exitosamente.'
            )

            return redirect('detalle_marca', pk=marca_temp.pk)

        messages.error(request, 'Por favor, corrija los errores en el formulario.')

    else:
        form = MarcaSenalForm(instance=marca)

    return render(request, 'app_sigrams/marcas/form.html', {
        'form': form,
        'titulo': f'Editar Marca #{marca.numero_orden}',
        'imagenes_predefinidas': ImagenMarcaPredefinida.objects.filter(activa=True),
        'marca': marca
    })



class ListaMarcasView(ListView):
    model = MarcaSenal
    template_name = 'app_sigrams/marcas/lista.html'
    context_object_name = 'marcas'
    paginate_by = 20

class NuevaMarcaView(CreateView):
    model = MarcaSenal
    form_class = MarcaSenalForm
    template_name = 'app_sigrams/marcas/form.html'
    success_url = reverse_lazy('lista_marcas')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Marca y Señal'
        context['imagenes_predefinidas'] = ImagenMarcaPredefinida.objects.filter(activa=True)
        return context

class EditarMarcaView(UpdateView):
    model = MarcaSenal
    form_class = MarcaSenalForm
    template_name = 'app_sigrams/marcas/form.html'
    success_url = reverse_lazy('lista_marcas')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Marca y Señal #{self.object.numero_orden}'
        context['imagenes_predefinidas'] = ImagenMarcaPredefinida.objects.filter(activa=True)
        return context

class DetalleMarcaView(DetailView):
    model = MarcaSenal
    template_name = 'app_sigrams/marcas/detalle.html'
    context_object_name = 'marca'

# Vista para cargar campos según el productor seleccionado (AJAX)
def cargar_campos(request):
    productor_id = request.GET.get('productor_id')
    campos = Campo.objects.filter(productor_id=productor_id).order_by('nombre')
    return render(request, 'marcas/campo_dropdown_options.html', {'campos': campos})





# ============================================================================
# VISTAS PARA SOLICITUDES
# ============================================================================

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count, Case, When, IntegerField
from django.utils import timezone
from datetime import timedelta
from app_registros.forms import SolicitudForm, SolicitudRevisionForm

@login_required
def lista_solicitudes(request):
    """Lista todas las solicitudes con filtros avanzados"""
    # Obtener todas las solicitudes
    solicitudes = Solicitud.objects.select_related(
        'productor', 'marca_senal', 'solicitante', 'revisor', 'aprobador'
    ).all().order_by('-fecha_solicitud')
    
    # Estadísticas
    total_solicitudes = solicitudes.count()
    solicitudes_pendientes = solicitudes.filter(estado='PENDIENTE').count()
    solicitudes_en_revision = solicitudes.filter(estado='EN_REVISION').count()
    solicitudes_vencidas = solicitudes.filter(
        Q(fecha_vencimiento__lt=timezone.now()) & 
        ~Q(estado__in=['APROBADO', 'RECHAZADO'])
    ).count()
    
    # Filtros
    estado = request.GET.get('estado', '')
    tipo_tramite = request.GET.get('tipo_tramite', '')
    prioridad = request.GET.get('prioridad', '')
    productor_id = request.GET.get('productor', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    
    if tipo_tramite:
        solicitudes = solicitudes.filter(tipo_tramite=tipo_tramite)
    
    if prioridad:
        solicitudes = solicitudes.filter(prioridad=prioridad)
    
    if productor_id:
        solicitudes = solicitudes.filter(productor_id=productor_id)
    
    if fecha_inicio:
        solicitudes = solicitudes.filter(fecha_solicitud__gte=fecha_inicio)
    
    if fecha_fin:
        solicitudes = solicitudes.filter(fecha_solicitud__lte=fecha_fin)
    
    # Paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(solicitudes, 20)  # 20 items por página
    
    try:
        solicitudes_pagina = paginator.page(page)
    except PageNotAnInteger:
        solicitudes_pagina = paginator.page(1)
    except EmptyPage:
        solicitudes_pagina = paginator.page(paginator.num_pages)
    
    context = {
        'solicitudes': solicitudes_pagina,
        'total_solicitudes': total_solicitudes,
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_en_revision': solicitudes_en_revision,
        'solicitudes_vencidas': solicitudes_vencidas,
        'estado_filtro': estado,
        'tipo_tramite_filtro': tipo_tramite,
        'prioridad_filtro': prioridad,
        'productor_filtro': productor_id,
        'fecha_inicio_filtro': fecha_inicio,
        'fecha_fin_filtro': fecha_fin,
        'productores': Productor.objects.all(),
    }
    return render(request, 'app_sigrams/solicitudes/lista.html', context)

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages

from app_registros.models import Solicitud, UserProfile
from app_registros.forms import DocumentoSolicitudFormSet


@login_required
def detalle_solicitud(request, pk):
    """Detalle de una solicitud específica con gestión de documentos"""

    solicitud = get_object_or_404(
        Solicitud.objects.select_related(
            'productor',
            'marca_senal',
            'solicitante',
            'revisor',
            'aprobador'
        ),
        pk=pk
    )

    # Perfil del usuario (lo mantenemos de tu código)

    default_rol = 'admin' if request.user.is_superuser else 'empleado'
    user_profile, created = UserProfile.objects.get_or_create(
    user=request.user,
    defaults={'rol': default_rol}
    )
    if created:
        messages.info(request, 'Perfil de usuario creado automáticamente.')

    # =========================
    # FORMSET DE DOCUMENTOS
    # =========================
    documento_formset = DocumentoSolicitudFormSet(
        instance=solicitud,
        queryset=solicitud.documentos.all().order_by('-fecha_subida')
    )

    # =========================
    # PROCESAR POST (DOCUMENTOS)
    # =========================
    if request.method == 'POST' and 'documentos' in request.POST:
        documento_formset = DocumentoSolicitudFormSet(
            request.POST,
            request.FILES,
            instance=solicitud
        )

        if documento_formset.is_valid():
            instancias = documento_formset.save(commit=False)

            for instancia in instancias:
                if instancia.archivo:
                    instancia.usuario_subida = request.user
                    instancia.save()

            documento_formset.save()
            messages.success(request, 'Documentos actualizados correctamente.')
            return redirect('detalle_solicitud', pk=pk)

    # =========================
    # CONTEXTO
    # =========================
    context = {
        'solicitud': solicitud,
        'user_profile': user_profile,
        'documento_formset': documento_formset,
        'roles_permitidos_documentos': ['admin', 'empleado'],
        'estados_para_devolver': ['EN_REVISION', 'OBSERVADO'],
    }

    return render(request, 'app_sigrams/solicitudes/detalle.html', context)

from app_registros.models import ChangeLog
@login_required
def nueva_solicitud(request):
    """Crear nueva solicitud"""
    if request.method == 'POST':
        form = SolicitudForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            solicitud = form.save()
            
            # Crear registro en el log de cambios
            ChangeLog.objects.create(
                user=request.user,
                modelo='Solicitud',
                objeto_id=solicitud.id,
                accion='CREACION',
                snapshot={
                    'tipo_tramite': solicitud.tipo_tramite,
                    'productor': str(solicitud.productor),
                    'estado': solicitud.estado
                }
            )
            
            messages.success(request, f'Solicitud #{solicitud.id} creada exitosamente.')
            return redirect('detalle_solicitud', pk=solicitud.pk)
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = SolicitudForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Nueva Solicitud',
        'productores': Productor.objects.all(),
    }
    return render(request, 'app_sigrams/solicitudes/form.html', context)

@login_required
def editar_solicitud(request, pk):
    """Editar solicitud existente"""
    solicitud = get_object_or_404(Solicitud, pk=pk)
    
    # Verificar que el usuario pueda editar (solo el solicitante o admin)
    
    default_rol = 'admin' if request.user.is_superuser else 'empleado'
    user_profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={'rol': default_rol}
    )
    if created:
        messages.info(request, 'Perfil de usuario creado automáticamente.')
    
    if not (solicitud.solicitante == request.user or user_profile.rol == 'admin'):
        messages.error(request, 'No tiene permisos para editar esta solicitud.')
        return redirect('detalle_solicitud', pk=pk)
    
    if request.method == 'POST':
        form = SolicitudForm(request.POST, request.FILES, instance=solicitud, user=request.user)
        if form.is_valid():
            solicitud = form.save()
            
            ChangeLog.objects.create(
                user=request.user,
                modelo='Solicitud',
                objeto_id=solicitud.id,
                accion='MODIFICACION',
                snapshot={
                    'tipo_tramite': solicitud.tipo_tramite,
                    'estado': solicitud.estado
                }
            )
            
            messages.success(request, f'Solicitud #{solicitud.id} actualizada exitosamente.')
            return redirect('detalle_solicitud', pk=solicitud.pk)
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = SolicitudForm(instance=solicitud, user=request.user)
    
    context = {
        'form': form,
        'titulo': f'Editar Solicitud #{solicitud.id}',
        'solicitud': solicitud,
    }
    return render(request, 'app_sigrams/solicitudes/form.html', context)

@login_required
def cambiar_estado_solicitud(request, pk, accion):
    """Cambiar estado de una solicitud"""
    solicitud = get_object_or_404(Solicitud, pk=pk)
    
    default_rol = 'admin' if request.user.is_superuser else 'empleado'
    user_profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={'rol': default_rol}
    )
    if created:
        messages.info(request, 'Perfil de usuario creado automáticamente.')
    # Import acá para evitar dependencias circulares
    from app_notificaciones.models import Notificacion

    # Verificar permisos según la acción
    if accion == 'revisar' and user_profile.rol in ['admin', 'empleado']:
        solicitud.estado = 'EN_REVISION'
        solicitud.revisor = request.user
        solicitud.fecha_revision = timezone.now()
        mensaje = f'Solicitud #{solicitud.id} puesta en revisión.'

        # Notificación
        if solicitud.solicitante:
            Notificacion.crear_notificacion(
                usuario=solicitud.solicitante,
                titulo='Solicitud en revisión',
                mensaje=f'Tu solicitud #{solicitud.id} está siendo revisada.',
                tipo='info',
                contenido_relacionado=solicitud,
                url=f'/solicitudes/{solicitud.id}/'
            )

    elif accion == 'aprobar' and user_profile.rol in ['admin', 'empleado']:
        solicitud.estado = 'APROBADO'
        solicitud.aprobador = request.user
        solicitud.fecha_resolucion = timezone.now()
        mensaje = f'Solicitud #{solicitud.id} aprobada.'

        # Notificación
        if solicitud.solicitante:
            Notificacion.crear_notificacion(
                usuario=solicitud.solicitante,
                titulo='¡Solicitud Aprobada!',
                mensaje=f'Tu solicitud #{solicitud.id} ha sido aprobada.',
                tipo='exito',
                contenido_relacionado=solicitud,
                url=f'/solicitudes/{solicitud.id}/'
            )

    elif accion == 'rechazar' and user_profile.rol in ['admin', 'empleado']:
        solicitud.estado = 'RECHAZADO'
        solicitud.fecha_resolucion = timezone.now()
        mensaje = f'Solicitud #{solicitud.id} rechazada.'

        # Notificación
        if solicitud.solicitante:
            Notificacion.crear_notificacion(
                usuario=solicitud.solicitante,
                titulo='Solicitud Rechazada',
                mensaje=f'Tu solicitud #{solicitud.id} ha sido rechazada.',
                tipo='error',
                contenido_relacionado=solicitud,
                url=f'/solicitudes/{solicitud.id}/'
            )

    elif accion == 'observar' and user_profile.rol in ['admin', 'empleado']:
        solicitud.estado = 'OBSERVADO'
        mensaje = f'Solicitud #{solicitud.id} marcada como observada.'

        # Notificación
        if solicitud.solicitante:
            Notificacion.crear_notificacion(
                usuario=solicitud.solicitante,
                titulo='Solicitud Observada',
                mensaje=f'Tu solicitud #{solicitud.id} requiere observaciones.',
                tipo='alerta',
                contenido_relacionado=solicitud,
                url=f'/solicitudes/{solicitud.id}/'
            )

    elif accion == 'pendiente' and user_profile.rol in ['admin', 'empleado']:
        solicitud.estado = 'PENDIENTE'
        mensaje = f'Solicitud #{solicitud.id} devuelta a pendiente.'

    else:
        messages.error(request, 'No tiene permisos para realizar esta acción.')
        return redirect('detalle_solicitud', pk=pk)

    solicitud.save()

    ChangeLog.objects.create(
        user=request.user,
        modelo='Solicitud',
        objeto_id=solicitud.id,
        accion=accion.upper(),
        snapshot={'estado': solicitud.estado}
    )

    messages.success(request, mensaje)
    return redirect('detalle_solicitud', pk=pk)


@login_required
def revision_solicitud(request, pk):
    """Formulario de revisión de solicitud"""
    solicitud = get_object_or_404(Solicitud, pk=pk)
    
    default_rol = 'admin' if request.user.is_superuser else 'empleado'
    user_profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={'rol': default_rol}
    )
    if created:
        messages.info(request, 'Perfil de usuario creado automáticamente.')
    
    # Verificar permisos
    if user_profile.rol not in ['admin', 'empleado']:
        messages.error(request, 'No tiene permisos para revisar solicitudes.')
        return redirect('detalle_solicitud', pk=pk)
    
    if request.method == 'POST':
        form = SolicitudRevisionForm(request.POST, instance=solicitud)
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.revisor = request.user
            solicitud.fecha_revision = timezone.now()
            solicitud.save()
            
            ChangeLog.objects.create(
                user=request.user,
                modelo='Solicitud',
                objeto_id=solicitud.id,
                accion='REVISION',
                snapshot={'estado': solicitud.estado}
            )
            
            messages.success(request, f'Revisión de solicitud #{solicitud.id} guardada.')
            return redirect('detalle_solicitud', pk=pk)
    else:
        form = SolicitudRevisionForm(instance=solicitud)
    
    context = {
        'form': form,
        'solicitud': solicitud,
        'titulo': f'Revisión de Solicitud #{solicitud.id}',
    }
    return render(request, 'app_sigrams/solicitudes/revision.html', context)

@login_required
def mis_solicitudes(request):
    """Lista las solicitudes del usuario actual"""
    solicitudes = Solicitud.objects.filter(solicitante=request.user).order_by('-fecha_solicitud')
    
    context = {
        'solicitudes': solicitudes,
        'titulo': 'Mis Solicitudes',
    }
    return render(request, 'app_sigrams/solicitudes/mis_solicitudes.html', context)

@login_required
def dashboard_solicitudes(request):
    """Dashboard con estadísticas de solicitudes"""
    from django.db.models import Count, Q
    
    # Estadísticas generales
    total_solicitudes = Solicitud.objects.count()
    solicitudes_pendientes = Solicitud.objects.filter(estado='PENDIENTE').count()
    solicitudes_en_revision = Solicitud.objects.filter(estado='EN_REVISION').count()
    solicitudes_aprobadas = Solicitud.objects.filter(estado='APROBADO').count()
    
    # Solicitudes por tipo de trámite
    solicitudes_por_tipo = Solicitud.objects.values('tipo_tramite').annotate(
        total=Count('id')
    ).order_by('tipo_tramite')
    
    # Solicitudes por prioridad
    solicitudes_por_prioridad = Solicitud.objects.values('prioridad').annotate(
        total=Count('id')
    ).order_by('prioridad')
    
    # Solicitudes vencidas
    solicitudes_vencidas = Solicitud.objects.filter(
        Q(fecha_vencimiento__lt=timezone.now()) & 
        ~Q(estado__in=['APROBADO', 'RECHAZADO'])
    ).count()
    
    # Últimas solicitudes
    ultimas_solicitudes = Solicitud.objects.select_related('productor', 'solicitante').order_by('-fecha_solicitud')[:10]
    
    context = {
        'total_solicitudes': total_solicitudes,
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_en_revision': solicitudes_en_revision,
        'solicitudes_aprobadas': solicitudes_aprobadas,
        'solicitudes_vencidas': solicitudes_vencidas,
        'solicitudes_por_tipo': list(solicitudes_por_tipo),
        'solicitudes_por_prioridad': list(solicitudes_por_prioridad),
        'ultimas_solicitudes': ultimas_solicitudes,
    }
    return render(request, 'app_sigrams/solicitudes/dashboard.html', context)



@login_required
def api_productores_geojson(request):
    """API para obtener productores en formato GeoJSON"""
    productores = Productor.objects.exclude(latitud__isnull=True).exclude(longitud__isnull=True)
    
    features = []
    for productor in productores:
        features.append({
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [float(productor.longitud), float(productor.latitud)]
            },
            "properties": {
                "id": productor.id,
                "nombre": productor.nombre_completo,
                "dni": productor.dni,
                "estado": productor.estado,
                "area_hectareas": float(productor.area_hectareas) if productor.area_hectareas else 0,
                "detalle_url": f"/productores/{productor.id}/"
            }
        })
    
    geojson = {
        "type": "FeatureCollection",
        "features": features
    }
    
    return JsonResponse(geojson)


# app_sigrams/views.py
class MapaProductoresView(LoginRequiredMixin, TemplateView):
    template_name = 'mapa/mapa.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener productores con coordenadas
        productores = Productor.objects.exclude(latitud__isnull=True).exclude(longitud__isnull=True)
        
        # Preparar datos para el mapa
        puntos_mapa = []
        for p in productores:
            if p.latitud and p.longitud:
                puntos_mapa.append({
                    'id': p.id,
                    'nombre': p.nombre_completo,
                    'dni': p.dni,
                    'localidad': p.localidad,
                    'estado': p.get_estado_display(),
                    'lat': float(p.latitud),
                    'lng': float(p.longitud),
                    'color': 'green' if p.estado == 'REGISTRADO' else 'orange'
                })
        
        context['puntos_mapa'] = puntos_mapa
        context['total_puntos'] = len(puntos_mapa)
        
        return context

def get_puntos_mapa_json(request):
    """API para obtener puntos del mapa en formato GeoJSON"""
    productores = Productor.objects.exclude(latitud__isnull=True).exclude(longitud__isnull=True)
    
    features = []
    for p in productores:
        if p.latitud and p.longitud:
            feature = {
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': [float(p.longitud), float(p.latitud)]
                },
                'properties': {
                    'id': p.id,
                    'nombre': p.nombre_completo,
                    'dni': p.dni,
                    'localidad': p.localidad,
                    'estado': p.estado,
                    'estado_display': p.get_estado_display(),
                    'color': 'green' if p.estado == 'REGISTRADO' else 'orange'
                }
            }
            features.append(feature)
    
    geojson = {
        'type': 'FeatureCollection',
        'features': features
    }
    
    return JsonResponse(geojson)




from django.db.models.functions import TruncMonth
from django.db.models import Sum

@login_required
def reporte_ingresos(request):
    export = request.GET.get('export')

    # Filtros
    año = request.GET.get('año', datetime.datetime.now().year)
    mes = request.GET.get('mes', '')

    ingresos_qs = MarcaSenal.objects.filter(valor_sellado__isnull=False)

    if año:
        ingresos_qs = ingresos_qs.filter(fecha_creacion__year=año)
    if mes:
        ingresos_qs = ingresos_qs.filter(fecha_creacion__month=mes)

    # 🔥 SI ES EXPORTACIÓN
    if export == 'pdf':
        return exportar_ingresos_pdf(ingresos_qs)
    elif export == 'excel':
        return exportar_ingresos_excel(ingresos_qs)

    # --- Vista HTML normal ---
    ingresos_por_mes = ingresos_qs.annotate(
        mes=TruncMonth('fecha_creacion')
    ).values('mes').annotate(
        total=Sum('valor_sellado'),
        cantidad=Count('id')
    ).order_by('-mes')

    ingresos_detalle = ingresos_qs.select_related('productor').order_by('-fecha_creacion')[:100]

    total_general = ingresos_qs.aggregate(total=Sum('valor_sellado'))['total'] or 0
    cantidad_total = ingresos_qs.count()

    años_disponibles = MarcaSenal.objects.dates('fecha_creacion', 'year')

    context = {
        'ingresos_por_mes': ingresos_por_mes,
        'ingresos_detalle': ingresos_detalle,
        'total_general': total_general,
        'cantidad_total': cantidad_total,
        'año_seleccionado': año,
        'mes_seleccionado': mes,
        'años_disponibles': años_disponibles,
    }

    return render(request, 'app_sigrams/reportes/ingresos.html', context)


def exportar_ingresos_pdf(queryset):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ingresos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Reporte de Ingresos - SIGRAMS", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(
            f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles['Normal']
        )
    )
    elements.append(Spacer(1, 24))

    datos = queryset.select_related('productor').order_by('-fecha_creacion')[:500]

    data = [['Fecha', 'N° Orden', 'Productor', 'Tipo Trámite', 'Valor']]

    for m in datos:
        data.append([
            m.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            str(m.numero_orden),
            m.productor.nombre_completo[:40],
            m.get_tipo_tramite_display(),
            f"${m.valor_sellado}"
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0A2E5A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


def exportar_ingresos_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingresos"

    headers = ['Fecha', 'N° Orden', 'Productor', 'Tipo Trámite', 'Valor']
    ws.append(headers)

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0A2E5A", end_color="0A2E5A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    datos = queryset.select_related('productor').order_by('-fecha_creacion')[:500]

    for m in datos:
        ws.append([
            m.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            m.numero_orden,
            m.productor.nombre_completo,
            m.get_tipo_tramite_display(),
            float(m.valor_sellado) if m.valor_sellado else 0
        ])

    # Ajustar ancho columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ingresos.xlsx"'

    wb.save(response)
    return response


def test_marcas_view(request):
    """Vista temporal para testear la obtención de marcas"""
    from app_registros.models import Productor, MarcaSenal
    
    # Listar todos los productores con sus marcas
    productores = Productor.objects.all().prefetch_related('marcas_senales')
    
    result = []
    for productor in productores:
        result.append({
            'id': productor.id,
            'nombre': productor.nombre_completo,
            'dni': productor.dni,
            'total_marcas': productor.marcas_senales.count(),
            'marcas': [
                {
                    'id': m.id,
                    'numero_orden': m.numero_orden,
                    'descripcion': m.descripcion_marca[:50] + '...' if len(m.descripcion_marca) > 50 else m.descripcion_marca
                }
                for m in productor.marcas_senales.all()[:5]  # Mostrar solo 5
            ]
        })
    
    return JsonResponse(result, safe=False)    


@login_required
def eliminar_documento(request, documento_id):
    """Eliminar un documento adjunto"""
    documento = get_object_or_404(DocumentoSolicitud, id=documento_id)
    
    # Verificar permisos
    if documento.usuario_subida != request.user and not UserProfile.objects.get(user=request.user).rol == 'admin':
        messages.error(request, 'No tiene permisos para eliminar este documento.')
        return redirect('detalle_solicitud', pk=documento.solicitud.pk)
    
    # Eliminar archivo físico
    documento.archivo.delete(save=False)
    
    # Eliminar registro
    solicitud_pk = documento.solicitud.pk
    documento.delete()
    
    messages.success(request, 'Documento eliminado correctamente.')
    return redirect('detalle_solicitud', pk=solicitud_pk)



from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================
# EXPORTACIONES PDF
# ============================================

@login_required
def reporte_productores_pdf(request):
    # (Ya lo tienes implementado como ReporteProductoresPDFView)
    # Si quieres, puedes convertir esa clase en función o mantenerla.
    # Por simplicidad, usaré la clase existente:
    return ReporteProductoresPDFView.as_view()(request)

@login_required
def reporte_marcas_pdf(request):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    import datetime

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="marcas.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Reporte de Marcas y Señales - SIGRAMS", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 24))
    
    marcas = MarcaSenal.objects.select_related('productor').all().order_by('-fecha_inscripcion')[:500]
    data = [['N° Orden', 'Productor', 'Tipo Trámite', 'Estado', 'Fecha Inscripción', 'Total Ganado']]
    for m in marcas:
        data.append([
            str(m.numero_orden),
            m.productor.nombre_completo,
            m.get_tipo_tramite_display(),
            m.get_estado_display(),
            m.fecha_inscripcion.strftime('%d/%m/%Y'),
            str(m.total_ganado)
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0A2E5A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response

@login_required
def reporte_solicitudes_pdf(request):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    import datetime

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="solicitudes.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Reporte de Solicitudes - SIGRAMS", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 24))
    
    solicitudes = Solicitud.objects.select_related('productor', 'solicitante').all().order_by('-fecha_solicitud')[:500]
    data = [['ID', 'Productor', 'Tipo Trámite', 'Estado', 'Prioridad', 'Fecha Solicitud']]
    for s in solicitudes:
        data.append([
            str(s.id),
            s.productor.nombre_completo,
            s.get_tipo_tramite_display(),
            s.get_estado_display(),
            s.get_prioridad_display(),
            s.fecha_solicitud.strftime('%d/%m/%Y')
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0A2E5A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response

# ============================================
# EXPORTACIONES EXCEL
# ============================================

@login_required
def reporte_productores_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productores"
    
    headers = ['Apellido', 'Nombre', 'DNI', 'CUIT', 'Localidad', 'Departamento', 'Estado', 'Fecha Registro']
    ws.append(headers)
    
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0A2E5A", end_color="0A2E5A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    productores = Productor.objects.all().order_by('apellido', 'nombre')
    for p in productores:
        ws.append([
            p.apellido,
            p.nombre,
            p.dni,
            p.cuit or '',
            p.localidad or '',
            p.departamento or '',
            p.get_estado_display(),
            p.fecha_registro.strftime('%d/%m/%Y')
        ])
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="productores.xlsx"'
    wb.save(response)
    return response

@login_required
def reporte_marcas_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Marcas"
    
    headers = ['N° Orden', 'Productor', 'Tipo Trámite', 'Estado', 'Fecha Inscripción', 'Total Ganado']
    ws.append(headers)
    
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0A2E5A", end_color="0A2E5A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    marcas = MarcaSenal.objects.select_related('productor').all().order_by('-fecha_inscripcion')
    for m in marcas:
        ws.append([
            m.numero_orden,
            m.productor.nombre_completo,
            m.get_tipo_tramite_display(),
            m.get_estado_display(),
            m.fecha_inscripcion.strftime('%d/%m/%Y'),
            m.total_ganado
        ])
    
    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="marcas.xlsx"'
    wb.save(response)
    return response

@login_required
def reporte_solicitudes_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"
    
    headers = ['ID', 'Productor', 'Tipo Trámite', 'Estado', 'Prioridad', 'Fecha Solicitud']
    ws.append(headers)
    
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0A2E5A", end_color="0A2E5A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    solicitudes = Solicitud.objects.select_related('productor').all().order_by('-fecha_solicitud')
    for s in solicitudes:
        ws.append([
            s.id,
            s.productor.nombre_completo,
            s.get_tipo_tramite_display(),
            s.get_estado_display(),
            s.get_prioridad_display(),
            s.fecha_solicitud.strftime('%d/%m/%Y')
        ])
    
    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="solicitudes.xlsx"'
    wb.save(response)
    return response


@login_required
def reporte_usuarios_pdf(request):
    # Verificar admin
    user_profile = UserProfile.objects.get(user=request.user)
    if user_profile.rol != 'admin':
        messages.error(request, 'No tiene permisos.')
        return redirect('home')

    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    import datetime

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="usuarios.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Reporte de Usuarios - SIGRAMS", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 24))

    usuarios = User.objects.select_related('userprofile').all().order_by('username')
    data = [['Usuario', 'Email', 'Rol', 'Fecha registro', 'Último acceso']]
    for u in usuarios:
        try:
            rol = u.userprofile.get_rol_display()
        except:
            rol = 'Sin perfil'
        data.append([
            u.username,
            u.email or '',
            rol,
            u.date_joined.strftime('%d/%m/%Y'),
            u.last_login.strftime('%d/%m/%Y %H:%M') if u.last_login else ''
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0A2E5A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

@login_required
def reporte_usuarios_excel(request):
    user_profile = UserProfile.objects.get(user=request.user)
    if user_profile.rol != 'admin':
        messages.error(request, 'No tiene permisos.')
        return redirect('home')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = ['Usuario', 'Email', 'Rol', 'Fecha registro', 'Último acceso']
    ws.append(headers)

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0A2E5A", end_color="0A2E5A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    usuarios = User.objects.select_related('userprofile').all().order_by('username')
    for u in usuarios:
        try:
            rol = u.userprofile.get_rol_display()
        except:
            rol = 'Sin perfil'
        ws.append([
            u.username,
            u.email or '',
            rol,
            u.date_joined.strftime('%d/%m/%Y'),
            u.last_login.strftime('%d/%m/%Y %H:%M') if u.last_login else ''
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# VISTAS DE BÚSQUEDA Y VERIFICACIÓN DE IMÁGENES
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def buscar_imagen_similar(request):
    if request.method == 'GET':
        return render(request, 'app_sigrams/marcas/buscar_imagen.html')

    try:
        from app_registros.image_similarity import calcular_hash_desde_bytes, buscar_en_todo
    except ImportError as e:
        return JsonResponse({'ok': False, 'error': f'Módulo no disponible: {e}'}, status=500)

    imagen = request.FILES.get('imagen')
    if not imagen:
        return JsonResponse({'ok': False, 'error': 'No se recibió ninguna imagen.'}, status=400)

    try:
        img_bytes = imagen.read()
        hash_nuevo = calcular_hash_desde_bytes(img_bytes)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Error al procesar imagen: {e}'}, status=422)

    if not hash_nuevo:
        return JsonResponse({'ok': False, 'error': 'No se pudo calcular el hash de la imagen.'}, status=422)

    try:
        resultados = buscar_en_todo(hash_nuevo)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Error en la búsqueda: {e}'}, status=500)

    for marca in resultados['marcas']:
        try:
            marca['url_detalle'] = reverse('detalle_marca', kwargs={'pk': marca['id']})
        except Exception:
            marca['url_detalle'] = f"/marcas/{marca['id']}/"

    total = len(resultados['marcas']) + len(resultados['predefinidas'])

    return JsonResponse({'ok': True, 'hash': hash_nuevo, 'resultados': resultados, 'total': total})


@login_required
def verificar_imagen_ajax(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    try:
        from app_registros.image_similarity import calcular_hash_desde_bytes, buscar_en_todo
    except ImportError as e:
        return JsonResponse({'ok': False, 'error': f'Módulo no disponible: {e}'}, status=500)

    imagen = request.FILES.get('imagen')
    if not imagen:
        return JsonResponse({'ok': False, 'error': 'No se recibió imagen.'}, status=400)

    excluir_id = request.POST.get('excluir_id')
    try:
        excluir_id = int(excluir_id) if excluir_id else None
    except (ValueError, TypeError):
        excluir_id = None

    try:
        img_bytes = imagen.read()
        hash_nuevo = calcular_hash_desde_bytes(img_bytes)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Error al procesar imagen: {e}'}, status=422)

    if not hash_nuevo:
        return JsonResponse({'ok': False, 'error': 'No se pudo procesar la imagen.'}, status=422)

    try:
        resultados = buscar_en_todo(hash_nuevo, excluir_marca_id=excluir_id)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Error en la búsqueda: {e}'}, status=500)

    total = len(resultados['marcas']) + len(resultados['predefinidas'])
    es_duplicada = total > 0

    mensaje = None
    if resultados['marcas']:
        mejor = resultados['marcas'][0]
        mensaje = (f'⚠️ Esta imagen ya está registrada para "{mejor["productor"]}" '
                   f'(Marca #{mejor["numero_orden"]}).')
    elif resultados['predefinidas']:
        mejor = resultados['predefinidas'][0]
        mensaje = f'⚠️ Esta imagen ya existe como predefinida: "{mejor["nombre"]}".'

    return JsonResponse({'ok': True, 'es_duplicada': es_duplicada, 'mensaje': mensaje, 'duplicados': resultados})


@login_required
def buscar_marca_por_nombre(request):
    """
    GET /ajax/buscar-marca-nombre/?q=texto
    Busca marcas e imágenes predefinidas por nombre o descripción.
    """
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'ok': False, 'error': 'Ingresá al menos 2 caracteres.'})

    # Buscar en MarcaSenal por descripcion_marca
    marcas_qs = MarcaSenal.objects.filter(
        Q(descripcion_marca__icontains=q) |
        Q(imagenes_predefinidas__nombre__icontains=q)
    ).select_related('productor').distinct()[:20]

    marcas = []
    for m in marcas_qs:
        try:
            url = reverse('detalle_marca', kwargs={'pk': m.pk})
        except Exception:
            url = f"/marcas/{m.pk}/"
        marcas.append({
            'id': m.pk,
            'numero_orden': m.numero_orden,
            'productor': str(m.productor),
            'productor_id': m.productor.pk,
            'descripcion': m.descripcion_marca[:100],
            'url_detalle': url,
            'imagen_url': m.imagen_marca.url if m.imagen_marca else None,
        })

    # Buscar en ImagenMarcaPredefinida por nombre
    predefinidas_qs = ImagenMarcaPredefinida.objects.filter(
        nombre__icontains=q, activa=True
    )[:10]

    predefinidas = [{
        'id': p.pk,
        'nombre': p.nombre,
        'tipo_marca': p.get_tipo_marca_display(),
        'imagen_url': p.imagen.url if p.imagen else None,
    } for p in predefinidas_qs]

    total = len(marcas) + len(predefinidas)
    return JsonResponse({'ok': True, 'query': q, 'marcas': marcas, 'predefinidas': predefinidas, 'total': total})


